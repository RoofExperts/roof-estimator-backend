"""
Condition-Based Commercial Roofing Estimating Engine - FastAPI Routes

This module provides all API endpoints for:
- Managing roof conditions (CRUD) with nested materials
- Managing condition materials (CRUD per condition)
- Managing material templates (CRUD)
- Managing the cost database (CRUD)
- Calculating estimates (conditions-driven)
- Smart Build with auto-populated materials
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
from pydantic import BaseModel
import datetime
import csv
import io

from database import SessionLocal
from auth import get_current_user
from conditions_models import (
    RoofCondition, MaterialTemplate, EstimateLineItem, CostDatabaseItem,
    ConditionMaterial, CONDITION_TYPES
)
from models import SavedEstimate
from estimate_engine import calculate_estimate, get_estimate_summary
from estimate_engine import get_available_condition_types, get_materials_for_condition
from estimate_engine import _find_cost_item, _calculate_quantity
from condition_builder import smart_build_conditions, _populate_materials_for_condition
from takeoff_engine import generate_takeoff


# ============================================================================
# DEPENDENCY: GET DATABASE SESSION
# ============================================================================

def get_db():
    """Dependency to get database session."""
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ============================================================================
# PYDANTIC SCHEMAS
# ============================================================================

class RoofConditionCreate(BaseModel):
    """Schema for creating a roof condition."""
    condition_type: str
    description: Optional[str] = None
    measurement_value: float
    measurement_unit: str = "sqft"
    wind_zone: Optional[str] = None
    flashing_height: Optional[float] = None
    fastener_spacing: Optional[int] = None


class RoofConditionUpdate(BaseModel):
    """Schema for updating a roof condition."""
    condition_type: Optional[str] = None
    description: Optional[str] = None
    measurement_value: Optional[float] = None
    measurement_unit: Optional[str] = None
    wind_zone: Optional[str] = None
    flashing_height: Optional[float] = None
    fastener_spacing: Optional[int] = None


class ConditionMaterialCreate(BaseModel):
    """Schema for adding a material to a condition."""
    material_name: str
    material_category: str
    unit: str
    coverage_rate: float
    waste_factor: float = 0.10
    calc_type: Optional[str] = None
    is_included: bool = True
    notes: Optional[str] = None
    cost_database_item_id: Optional[int] = None


class ConditionMaterialUpdate(BaseModel):
    """Schema for updating a condition material."""
    material_name: Optional[str] = None
    coverage_rate: Optional[float] = None
    waste_factor: Optional[float] = None
    calc_type: Optional[str] = None
    is_included: Optional[bool] = None
    override_quantity: Optional[float] = None
    notes: Optional[str] = None
    sort_order: Optional[int] = None
    cost_database_item_id: Optional[int] = None


class ConditionMaterialResponse(BaseModel):
    id: int
    condition_id: int
    material_template_id: Optional[int]
    material_name: str
    material_category: str
    unit: str
    coverage_rate: float
    waste_factor: float
    calc_type: Optional[str]
    is_included: bool
    override_quantity: Optional[float]
    notes: Optional[str]
    sort_order: int
    cost_database_item_id: Optional[int] = None

    class Config:
        from_attributes = True


class RoofConditionResponse(BaseModel):
    """Schema for roof condition response."""
    id: int
    project_id: int
    condition_type: str
    description: Optional[str]
    measurement_value: float
    measurement_unit: str
    wind_zone: Optional[str]
    flashing_height: Optional[float] = None
    fastener_spacing: Optional[int] = None
    created_at: datetime.datetime

    class Config:
        from_attributes = True


class ConditionWithMaterialsResponse(BaseModel):
    """Condition with its nested materials."""
    id: int
    project_id: int
    condition_type: str
    condition_label: str  # human-readable label
    description: Optional[str]
    measurement_value: float
    measurement_unit: str
    wind_zone: Optional[str]
    flashing_height: Optional[float] = None
    fastener_spacing: Optional[int] = None
    materials: List[ConditionMaterialResponse]

    class Config:
        from_attributes = True


class MaterialTemplateCreate(BaseModel):
    """Schema for creating a material template."""
    system_type: str = "common"
    condition_type: str
    material_name: str
    material_category: str
    unit: str
    coverage_rate: float
    waste_factor: float = 0.10
    is_active: bool = True


class MaterialTemplateUpdate(BaseModel):
    """Schema for updating a material template."""
    coverage_rate: Optional[float] = None
    waste_factor: Optional[float] = None
    is_active: Optional[bool] = None


class MaterialTemplateResponse(BaseModel):
    """Schema for material template response."""
    id: int
    system_type: str
    condition_type: str
    material_name: str
    material_category: str
    unit: str
    coverage_rate: float
    waste_factor: float
    is_active: bool
    created_at: datetime.datetime

    class Config:
        from_attributes = True


class CostDatabaseItemCreate(BaseModel):
    """Schema for creating a cost database item."""
    material_name: str
    manufacturer: Optional[str] = None
    material_category: str
    unit: str
    unit_cost: float
    labor_cost_per_unit: Optional[float] = None
    is_active: bool = True
    purchase_unit: Optional[str] = None
    units_per_purchase: Optional[float] = None
    product_name: Optional[str] = None


class CostDatabaseItemUpdate(BaseModel):
    """Schema for updating a cost database item."""
    unit_cost: Optional[float] = None
    labor_cost_per_unit: Optional[float] = None
    is_active: Optional[bool] = None
    purchase_unit: Optional[str] = None
    units_per_purchase: Optional[float] = None
    product_name: Optional[str] = None


class CostDatabaseItemResponse(BaseModel):
    """Schema for cost database item response."""
    id: int
    material_name: str
    manufacturer: Optional[str]
    material_category: str
    unit: str
    unit_cost: float
    labor_cost_per_unit: Optional[float]
    last_updated: datetime.datetime
    is_active: bool
    purchase_unit: Optional[str] = None
    units_per_purchase: Optional[float] = None
    product_name: Optional[str] = None

    class Config:
        from_attributes = True


class EstimateLineItemResponse(BaseModel):
    """Schema for estimate line item response."""
    id: int
    project_id: int
    condition_id: int
    material_name: str
    material_category: str
    quantity: float
    unit: str
    unit_cost: float
    total_cost: float
    notes: Optional[str]
    created_at: datetime.datetime

    class Config:
        from_attributes = True


# ============================================================================
# CREATE ROUTER
# ============================================================================

router = APIRouter(prefix="/api/v1", tags=["roofing-estimate"])


# ============================================================================
# ROOF CONDITION ENDPOINTS
# ============================================================================

@router.post("/projects/{project_id}/conditions", response_model=RoofConditionResponse)
def create_condition(
    project_id: int,
    condition: RoofConditionCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Add a new roof condition to a project and auto-populate materials from templates."""
    from models import Project

    roof_condition = RoofCondition(
        project_id=project_id,
        condition_type=condition.condition_type,
        description=condition.description,
        measurement_value=condition.measurement_value,
        measurement_unit=condition.measurement_unit,
        wind_zone=condition.wind_zone,
        flashing_height=condition.flashing_height,
        fastener_spacing=condition.fastener_spacing,
    )
    db.add(roof_condition)
    db.flush()  # Get the condition ID

    # ── Auto-populate materials from templates (Edge workflow) ──
    project = db.query(Project).filter(Project.id == project_id).first()
    system_type = (project.system_type or "TPO") if project else "TPO"
    org_id = current_user.get("org_id", 1)
    mat_count = _populate_materials_for_condition(roof_condition, system_type, org_id, db)
    print(f"[CreateCondition] Auto-populated {mat_count} materials for {condition.condition_type}")

    db.commit()
    db.refresh(roof_condition)
    return roof_condition


@router.get("/projects/{project_id}/conditions", response_model=List[RoofConditionResponse])
def list_conditions(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List all conditions for a project."""
    conditions = db.query(RoofCondition).filter(
        RoofCondition.project_id == project_id
    ).all()
    return conditions


@router.get("/projects/{project_id}/conditions-with-materials")
def list_conditions_with_materials(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    List all conditions for a project WITH their nested materials.
    Enriched with calculated quantities and cost data from the cost database.
    This is the primary endpoint for the Conditions accordion UI.
    """
    from models import Project

    project = db.query(Project).filter(Project.id == project_id).first()
    org_id = project.org_id if project else current_user["org_id"]

    conditions = db.query(RoofCondition).filter(
        RoofCondition.project_id == project_id
    ).all()

    result = []
    for c in conditions:
        materials = db.query(ConditionMaterial).filter(
            ConditionMaterial.condition_id == c.id
        ).order_by(ConditionMaterial.sort_order).all()

        ct_info = CONDITION_TYPES.get(c.condition_type, {"label": c.condition_type})

        enriched_materials = []
        for m in materials:
            # Calculate quantity using same formula as estimate engine
            qty_calculated = _calculate_quantity(c, m)

            # Look up cost data — prefer explicit link, fallback to fuzzy match
            cost_item = None
            if m.cost_database_item_id:
                cost_item = db.query(CostDatabaseItem).filter(
                    CostDatabaseItem.id == m.cost_database_item_id,
                    CostDatabaseItem.is_active == True,
                ).first()
            if not cost_item:
                cost_item = _find_cost_item(m.material_name, m.unit, org_id, db)

            unit_cost = cost_item.unit_cost if cost_item else 0.0
            labor_cost = cost_item.labor_cost_per_unit if cost_item else None
            extended_cost = round(qty_calculated * unit_cost, 2) if m.is_included else 0.0

            enriched_materials.append({
                "id": m.id,
                "condition_id": m.condition_id,
                "material_template_id": m.material_template_id,
                "material_name": m.material_name,
                "material_category": m.material_category,
                "unit": m.unit,
                "coverage_rate": m.coverage_rate,
                "waste_factor": m.waste_factor,
                "calc_type": m.calc_type,
                "is_included": m.is_included,
                "override_quantity": m.override_quantity,
                "notes": m.notes,
                "sort_order": m.sort_order,
                "cost_database_item_id": m.cost_database_item_id,
                # Enriched cost data
                "qty_calculated": round(qty_calculated, 2),
                "unit_cost": unit_cost,
                "labor_cost_per_unit": labor_cost,
                "extended_cost": extended_cost,
                "purchase_unit": cost_item.purchase_unit if cost_item else None,
                "units_per_purchase": cost_item.units_per_purchase if cost_item else None,
                "product_name": cost_item.product_name if cost_item else None,
                "cost_db_match": cost_item.material_name if cost_item else None,
            })

        result.append({
            "id": c.id,
            "project_id": c.project_id,
            "condition_type": c.condition_type,
            "condition_label": ct_info["label"],
            "description": c.description,
            "measurement_value": c.measurement_value,
            "measurement_unit": c.measurement_unit,
            "wind_zone": c.wind_zone,
            "flashing_height": c.flashing_height,
            "fastener_spacing": c.fastener_spacing,
            "materials": enriched_materials,
        })

    return result


@router.get("/conditions/{condition_id}", response_model=RoofConditionResponse)
def get_condition(
    condition_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get a specific roof condition by ID."""
    condition = db.query(RoofCondition).filter(
        RoofCondition.id == condition_id
    ).first()
    if not condition:
        raise HTTPException(status_code=404, detail="Condition not found")
    return condition


@router.put("/conditions/{condition_id}", response_model=RoofConditionResponse)
def update_condition(
    condition_id: int,
    condition: RoofConditionUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update an existing roof condition."""
    db_condition = db.query(RoofCondition).filter(
        RoofCondition.id == condition_id
    ).first()
    if not db_condition:
        raise HTTPException(status_code=404, detail="Condition not found")

    if condition.condition_type is not None:
        db_condition.condition_type = condition.condition_type
    if condition.description is not None:
        db_condition.description = condition.description
    if condition.measurement_value is not None:
        db_condition.measurement_value = condition.measurement_value
    if condition.measurement_unit is not None:
        db_condition.measurement_unit = condition.measurement_unit
    if condition.wind_zone is not None:
        db_condition.wind_zone = condition.wind_zone
    if condition.flashing_height is not None:
        db_condition.flashing_height = condition.flashing_height
    if condition.fastener_spacing is not None:
        db_condition.fastener_spacing = condition.fastener_spacing

    db.commit()
    db.refresh(db_condition)
    return db_condition


@router.delete("/conditions/{condition_id}")
def delete_condition(
    condition_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Delete a roof condition. Also deletes associated materials and line items."""
    db_condition = db.query(RoofCondition).filter(
        RoofCondition.id == condition_id
    ).first()
    if not db_condition:
        raise HTTPException(status_code=404, detail="Condition not found")

    db.query(EstimateLineItem).filter(
        EstimateLineItem.condition_id == condition_id
    ).delete()
    db.query(ConditionMaterial).filter(
        ConditionMaterial.condition_id == condition_id
    ).delete()

    db.delete(db_condition)
    db.commit()
    return {"message": "Condition deleted successfully"}


# ============================================================================
# CONDITION MATERIAL ENDPOINTS  (per-condition material management)
# ============================================================================

@router.get("/conditions/{condition_id}/materials", response_model=List[ConditionMaterialResponse])
def list_condition_materials(
    condition_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List all materials for a specific condition."""
    return db.query(ConditionMaterial).filter(
        ConditionMaterial.condition_id == condition_id
    ).order_by(ConditionMaterial.sort_order).all()


@router.post("/conditions/{condition_id}/materials", response_model=ConditionMaterialResponse)
def add_condition_material(
    condition_id: int,
    mat: ConditionMaterialCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Add a material to a condition (custom add)."""
    condition = db.query(RoofCondition).filter(RoofCondition.id == condition_id).first()
    if not condition:
        raise HTTPException(status_code=404, detail="Condition not found")

    # Get max sort_order for this condition
    max_order = db.query(ConditionMaterial).filter(
        ConditionMaterial.condition_id == condition_id
    ).count()

    cm = ConditionMaterial(
        condition_id=condition_id,
        material_name=mat.material_name,
        material_category=mat.material_category,
        unit=mat.unit,
        coverage_rate=mat.coverage_rate,
        waste_factor=mat.waste_factor,
        calc_type=mat.calc_type,
        is_included=mat.is_included,
        notes=mat.notes,
        sort_order=max_order,
        cost_database_item_id=mat.cost_database_item_id,
    )
    db.add(cm)
    db.commit()
    db.refresh(cm)
    return cm


@router.put("/condition-materials/{material_id}", response_model=ConditionMaterialResponse)
def update_condition_material(
    material_id: int,
    mat: ConditionMaterialUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update a condition material (coverage rate, toggle, override qty, etc.)."""
    cm = db.query(ConditionMaterial).filter(ConditionMaterial.id == material_id).first()
    if not cm:
        raise HTTPException(status_code=404, detail="Condition material not found")

    if mat.material_name is not None:
        cm.material_name = mat.material_name
    if mat.coverage_rate is not None:
        cm.coverage_rate = mat.coverage_rate
    if mat.waste_factor is not None:
        cm.waste_factor = mat.waste_factor
    if mat.calc_type is not None:
        cm.calc_type = mat.calc_type
    if mat.is_included is not None:
        cm.is_included = mat.is_included
    if mat.override_quantity is not None:
        cm.override_quantity = mat.override_quantity
    if mat.notes is not None:
        cm.notes = mat.notes
    if mat.sort_order is not None:
        cm.sort_order = mat.sort_order
    if mat.cost_database_item_id is not None:
        cm.cost_database_item_id = mat.cost_database_item_id

    db.commit()
    db.refresh(cm)
    return cm


@router.delete("/condition-materials/{material_id}")
def delete_condition_material(
    material_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Remove a material from a condition."""
    cm = db.query(ConditionMaterial).filter(ConditionMaterial.id == material_id).first()
    if not cm:
        raise HTTPException(status_code=404, detail="Condition material not found")
    db.delete(cm)
    db.commit()
    return {"message": "Material removed from condition"}


@router.post("/projects/{project_id}/populate-materials")
def populate_materials_for_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Auto-populate materials for ALL conditions in a project from MaterialTemplates.
    Only adds materials to conditions that don't already have them.
    Uses the project's system_type to pick the right templates.
    """
    from models import Project

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    system_type = (project.system_type or "TPO").upper()
    org_id = current_user["org_id"]

    conditions = db.query(RoofCondition).filter(
        RoofCondition.project_id == project_id
    ).all()

    total_added = 0

    for condition in conditions:
        # Skip if condition already has materials
        existing_count = db.query(ConditionMaterial).filter(
            ConditionMaterial.condition_id == condition.id
        ).count()
        if existing_count > 0:
            continue

        # Find matching templates (org-specific first, then global)
        templates = db.query(MaterialTemplate).filter(
            MaterialTemplate.condition_type == condition.condition_type,
            MaterialTemplate.is_active == True,
            or_(
                MaterialTemplate.system_type == system_type,
                MaterialTemplate.system_type == "common"
            ),
            or_(
                MaterialTemplate.org_id == org_id,
                MaterialTemplate.is_global == True
            )
        ).order_by(MaterialTemplate.material_category).all()

        for idx, tmpl in enumerate(templates):
            cm = ConditionMaterial(
                condition_id=condition.id,
                material_template_id=tmpl.id,
                material_name=tmpl.material_name,
                material_category=tmpl.material_category,
                unit=tmpl.unit,
                coverage_rate=tmpl.coverage_rate,
                waste_factor=tmpl.waste_factor,
                calc_type=tmpl.calc_type,
                is_included=True,
                sort_order=idx,
            )
            db.add(cm)
            total_added += 1

    db.commit()

    return {
        "message": f"Populated {total_added} materials across {len(conditions)} conditions",
        "materials_added": total_added,
        "conditions_count": len(conditions),
    }


# ============================================================================
# ESTIMATE CALCULATION ENDPOINTS
# ============================================================================

@router.post("/projects/{project_id}/calculate-estimate")
def calculate_project_estimate(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Calculate the complete estimate for a project."""
    result = calculate_estimate(project_id, db)
    if result.get("status") == "error":
        raise HTTPException(status_code=400, detail=result.get("message"))
    return result


@router.get("/projects/{project_id}/estimate")
def get_project_estimate(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get the full estimate summary for a project."""
    result = get_estimate_summary(project_id, db)
    if result.get("status") == "error":
        raise HTTPException(status_code=400, detail=result.get("message"))
    return result


# ============================================================================
# MATERIAL TEMPLATE ENDPOINTS
# ============================================================================

@router.post("/material-templates", response_model=MaterialTemplateResponse)
def create_material_template(
    template: MaterialTemplateCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Create a new material template."""
    db_template = MaterialTemplate(
        system_type=template.system_type,
        condition_type=template.condition_type,
        material_name=template.material_name,
        material_category=template.material_category,
        unit=template.unit,
        coverage_rate=template.coverage_rate,
        waste_factor=template.waste_factor,
        is_active=template.is_active,
        org_id=current_user["org_id"]
    )
    db.add(db_template)
    db.commit()
    db.refresh(db_template)
    return db_template


@router.get("/material-templates", response_model=List[MaterialTemplateResponse])
def list_material_templates(
    condition_type: Optional[str] = Query(None),
    system_type: Optional[str] = Query(None),
    is_active: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List material templates with optional filtering by condition_type and system_type."""
    query = db.query(MaterialTemplate).filter(
        MaterialTemplate.is_active == is_active,
        or_(MaterialTemplate.org_id == current_user["org_id"], MaterialTemplate.is_global == True)
    )
    if condition_type:
        query = query.filter(MaterialTemplate.condition_type == condition_type)
    if system_type:
        query = query.filter(or_(
            MaterialTemplate.system_type == system_type,
            MaterialTemplate.system_type == "common"
        ))
    return query.all()


@router.get("/material-templates/{template_id}", response_model=MaterialTemplateResponse)
def get_material_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get a specific material template by ID."""
    template = db.query(MaterialTemplate).filter(MaterialTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return template


@router.put("/material-templates/{template_id}", response_model=MaterialTemplateResponse)
def update_material_template(
    template_id: int,
    template: MaterialTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Update an existing material template."""
    db_template = db.query(MaterialTemplate).filter(
        MaterialTemplate.id == template_id,
        MaterialTemplate.org_id == current_user["org_id"]
    ).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="Template not found")
    if template.coverage_rate is not None:
        db_template.coverage_rate = template.coverage_rate
    if template.waste_factor is not None:
        db_template.waste_factor = template.waste_factor
    if template.is_active is not None:
        db_template.is_active = template.is_active
    db.commit()
    db.refresh(db_template)
    return db_template


@router.delete("/material-templates/{template_id}")
def delete_material_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Delete a material template (soft delete)."""
    db_template = db.query(MaterialTemplate).filter(
        MaterialTemplate.id == template_id,
        MaterialTemplate.org_id == current_user["org_id"]
    ).first()
    if not db_template:
        raise HTTPException(status_code=404, detail="Template not found")
    db_template.is_active = False
    db.commit()
    return {"message": "Template deleted successfully"}


# ============================================================================
# COST DATABASE ENDPOINTS
# ============================================================================

@router.post("/cost-database", response_model=CostDatabaseItemResponse)
def create_cost_item(
    item: CostDatabaseItemCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Add a new item to the cost database."""
    db_item = CostDatabaseItem(
        material_name=item.material_name,
        manufacturer=item.manufacturer,
        material_category=item.material_category,
        unit=item.unit,
        unit_cost=item.unit_cost,
        labor_cost_per_unit=item.labor_cost_per_unit,
        is_active=item.is_active,
        org_id=current_user["org_id"],
        purchase_unit=item.purchase_unit,
        units_per_purchase=item.units_per_purchase,
        product_name=item.product_name,
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item


@router.get("/cost-database", response_model=List[CostDatabaseItemResponse])
def list_cost_items(
    material_category: Optional[str] = Query(None),
    is_active: bool = Query(True),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """List cost database items with optional filtering."""
    query = db.query(CostDatabaseItem).filter(
        CostDatabaseItem.is_active == is_active,
        or_(CostDatabaseItem.org_id == current_user["org_id"], CostDatabaseItem.is_global == True)
    )
    if material_category:
        query = query.filter(CostDatabaseItem.material_category == material_category)
    return query.order_by(CostDatabaseItem.material_name).all()


@router.get("/cost-database/search")
def search_cost_database(
    q: str = Query("", description="Search term for material name"),
    category: Optional[str] = Query(None, description="Filter by material_category"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Search cost database items by name (partial match) and optional category.
    Used by the Conditions tab for Add Material and Product Swap flows.
    Returns matching items from org's cost DB + global fallbacks.
    """
    from sqlalchemy import func

    org_id = current_user["org_id"]
    query = db.query(CostDatabaseItem).filter(
        CostDatabaseItem.is_active == True,
        or_(CostDatabaseItem.org_id == org_id, CostDatabaseItem.is_global == True)
    )

    if category:
        query = query.filter(CostDatabaseItem.material_category == category)

    if q.strip():
        search_term = f"%{q.strip().lower()}%"
        query = query.filter(func.lower(CostDatabaseItem.material_name).like(search_term))

    results = query.order_by(CostDatabaseItem.material_name).limit(50).all()

    return [
        {
            "id": item.id,
            "material_name": item.material_name,
            "manufacturer": item.manufacturer,
            "material_category": item.material_category,
            "unit": item.unit,
            "unit_cost": item.unit_cost,
            "labor_cost_per_unit": item.labor_cost_per_unit,
            "purchase_unit": item.purchase_unit,
            "units_per_purchase": item.units_per_purchase,
            "product_name": item.product_name,
            "description": item.description,
        }
        for item in results
    ]


@router.post("/cost-database/upload-pricing")
def upload_pricing_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload vendor pricing from CSV or Excel file.
    """
    try:
        content = file.file.read()
        filename = file.filename.lower()

        rows = []

        if filename.endswith('.csv'):
            text_content = content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(text_content))
            rows = list(csv_reader)
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
            except ImportError:
                raise HTTPException(
                    status_code=400,
                    detail="Excel support requires openpyxl. Please upload a CSV file instead."
                )
            from openpyxl import load_workbook
            workbook = load_workbook(io.BytesIO(content))
            worksheet = workbook.active
            headers = [cell.value for cell in worksheet[1]]
            for row in worksheet.iter_rows(min_row=2, values_only=False):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx < len(headers):
                        row_data[headers[idx]] = cell.value
                rows.append(row_data)
        else:
            raise HTTPException(status_code=400, detail="File must be CSV or Excel (.csv, .xlsx, .xls)")

        if not rows:
            raise HTTPException(status_code=400, detail="File is empty")

        matched = 0
        updated = 0
        unmatched_items = []

        for row in rows:
            if not row:
                continue
            material_name = row.get('material_name') or row.get('Material Name')
            unit_cost = row.get('unit_cost') or row.get('Unit Cost')
            if not material_name or unit_cost is None:
                continue
            try:
                unit_cost = float(unit_cost)
            except (ValueError, TypeError):
                unmatched_items.append(str(material_name))
                continue
            labor_cost_per_unit = row.get('labor_cost_per_unit') or row.get('Labor Cost Per Unit')
            if labor_cost_per_unit is not None:
                try:
                    labor_cost_per_unit = float(labor_cost_per_unit)
                except (ValueError, TypeError):
                    labor_cost_per_unit = None
            matching_item = db.query(CostDatabaseItem).filter(
                CostDatabaseItem.material_name.ilike(f'%{material_name}%'),
                CostDatabaseItem.org_id == current_user["org_id"]
            ).first()
            if matching_item:
                matched += 1
                matching_item.unit_cost = unit_cost
                if labor_cost_per_unit is not None:
                    matching_item.labor_cost_per_unit = labor_cost_per_unit
                matching_item.last_updated = datetime.datetime.utcnow()
                updated += 1
            else:
                unmatched_items.append(str(material_name))

        db.commit()
        return {"total_rows": len(rows), "matched": matched, "updated": updated, "unmatched_items": unmatched_items}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")


@router.post("/cost-database/resync")
def resync_cost_database(
    update_pricing: bool = False,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Re-sync this org's cost database from platform defaults.
    - Adds any new global items that the org doesn't have yet.
    - Fills in missing purchase_unit data on existing items.
    - If update_pricing=True, also resets pricing to platform defaults.
    """
    from seed_data import resync_cost_items_for_org, update_global_purchase_units
    org_id = current_user["org_id"]
    # First ensure global items have purchase_unit data
    update_global_purchase_units(db)
    # Then resync org items from globals
    result = resync_cost_items_for_org(org_id, db, update_pricing=update_pricing)
    db.commit()
    return {
        "message": f"Resync complete: {result['added']} added, {result['updated']} updated",
        **result
    }


@router.get("/cost-database/{item_id}", response_model=CostDatabaseItemResponse)
def get_cost_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    item = db.query(CostDatabaseItem).filter(CostDatabaseItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Cost item not found")
    return item


@router.put("/cost-database/{item_id}", response_model=CostDatabaseItemResponse)
def update_cost_item(
    item_id: int,
    item: CostDatabaseItemUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    db_item = db.query(CostDatabaseItem).filter(
        CostDatabaseItem.id == item_id,
        CostDatabaseItem.org_id == current_user["org_id"]
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Cost item not found")
    if item.unit_cost is not None:
        db_item.unit_cost = item.unit_cost
    if item.labor_cost_per_unit is not None:
        db_item.labor_cost_per_unit = item.labor_cost_per_unit
    if item.is_active is not None:
        db_item.is_active = item.is_active
    if item.purchase_unit is not None:
        db_item.purchase_unit = item.purchase_unit
    if item.units_per_purchase is not None:
        db_item.units_per_purchase = item.units_per_purchase
    if item.product_name is not None:
        db_item.product_name = item.product_name
    db_item.last_updated = datetime.datetime.utcnow()
    db.commit()
    db.refresh(db_item)
    return db_item


@router.delete("/cost-database/{item_id}")
def delete_cost_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    db_item = db.query(CostDatabaseItem).filter(
        CostDatabaseItem.id == item_id,
        CostDatabaseItem.org_id == current_user["org_id"]
    ).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Cost item not found")
    db_item.is_active = False
    db_item.last_updated = datetime.datetime.utcnow()
    db.commit()
    return {"message": "Cost item deleted successfully"}


# ============================================================================
# CONDITION PRESET ENDPOINTS (Company Admin Portal)
# ============================================================================

@router.get("/material-templates/by-condition")
def get_templates_by_condition(
    condition_type: str = Query(...),
    system_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Get material templates grouped by condition type, ordered by sort_order.
    Returns both org-specific and global templates (org takes precedence).
    Used by the Company Admin Portal > Condition Presets tab.
    """
    org_id = current_user["org_id"]

    # Build query for org-specific templates first, fall back to global
    query = db.query(MaterialTemplate).filter(
        MaterialTemplate.condition_type == condition_type,
        MaterialTemplate.is_active == True,
        or_(
            MaterialTemplate.org_id == org_id,
            MaterialTemplate.is_global == True
        )
    )
    if system_type:
        query = query.filter(
            or_(MaterialTemplate.system_type == system_type,
                MaterialTemplate.system_type == "common")
        )

    templates = query.order_by(MaterialTemplate.sort_order, MaterialTemplate.id).all()

    # De-duplicate: if org has a custom version, skip the global one
    seen = set()
    result = []
    # Org-specific first
    for t in templates:
        if t.org_id == org_id:
            key = (t.system_type, t.material_name, t.material_category)
            seen.add(key)
            result.append(t)
    # Then global fallbacks
    for t in templates:
        if t.is_global and t.org_id is None:
            key = (t.system_type, t.material_name, t.material_category)
            if key not in seen:
                result.append(t)

    result.sort(key=lambda x: (x.sort_order or 0, x.id))

    return {
        "condition_type": condition_type,
        "system_type": system_type,
        "templates": [
            {
                "id": t.id,
                "system_type": t.system_type,
                "condition_type": t.condition_type,
                "material_name": t.material_name,
                "material_category": t.material_category,
                "unit": t.unit,
                "coverage_rate": t.coverage_rate,
                "waste_factor": t.waste_factor,
                "calc_type": t.calc_type,
                "sort_order": t.sort_order or 0,
                "is_optional": t.is_optional or False,
                "is_global": t.is_global,
                "is_active": t.is_active,
            }
            for t in result
        ]
    }


class ReorderRequest(BaseModel):
    template_ids: List[int]  # ordered list of template IDs in desired sort order


@router.put("/material-templates/reorder")
def reorder_templates(
    body: ReorderRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Bulk update sort_order for a list of material templates.
    Accepts an ordered list of template IDs. Sets sort_order = index * 10.
    """
    org_id = current_user["org_id"]
    updated = 0
    for idx, tid in enumerate(body.template_ids):
        t = db.query(MaterialTemplate).filter(MaterialTemplate.id == tid).first()
        if t and (t.org_id == org_id or t.is_global):
            # If it's a global template, clone it for this org first
            if t.is_global and t.org_id is None:
                new_t = MaterialTemplate(
                    system_type=t.system_type, condition_type=t.condition_type,
                    material_name=t.material_name, material_category=t.material_category,
                    unit=t.unit, coverage_rate=t.coverage_rate, waste_factor=t.waste_factor,
                    calc_type=t.calc_type, sort_order=(idx + 1) * 10,
                    is_optional=t.is_optional, is_active=True,
                    org_id=org_id, is_global=False
                )
                db.add(new_t)
            else:
                t.sort_order = (idx + 1) * 10
            updated += 1
    db.commit()
    return {"status": "success", "updated": updated}


class ResetConditionRequest(BaseModel):
    condition_type: str
    system_type: Optional[str] = None


@router.post("/material-templates/reset-condition")
def reset_condition_templates(
    body: ResetConditionRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Reset a condition type back to global defaults for this org.
    Deletes org-specific overrides so the global templates take effect.
    """
    org_id = current_user["org_id"]
    query = db.query(MaterialTemplate).filter(
        MaterialTemplate.org_id == org_id,
        MaterialTemplate.condition_type == body.condition_type
    )
    if body.system_type:
        query = query.filter(MaterialTemplate.system_type == body.system_type)

    deleted = query.delete()
    db.commit()
    return {"status": "success", "deleted": deleted, "message": f"Reset {deleted} org-specific templates. Global defaults will now apply."}


# ============================================================================
# REFERENCE DATA ENDPOINTS
# ============================================================================

@router.get("/reference/condition-types")
def get_condition_types(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get all available condition types with labels and default units."""
    return {
        "condition_types": [
            {"value": k, "label": v["label"], "default_unit": v["default_unit"]}
            for k, v in CONDITION_TYPES.items()
        ]
    }


@router.get("/reference/condition-types/{condition_type}/materials")
def get_condition_materials_ref(
    condition_type: str,
    system_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Get all materials available for a specific condition type, optionally filtered by system."""
    materials = get_materials_for_condition(condition_type, db, system_type=system_type)
    return {"condition_type": condition_type, "system_type": system_type, "materials": materials}


# ============================================================================
# SMART BUILD ENDPOINT
# ============================================================================

@router.post("/projects/{project_id}/smart-build-conditions")
def smart_build(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Intelligently build conditions from spec analysis + plan extractions,
    then auto-populate materials for each condition from templates.
    """
    result = smart_build_conditions(project_id, db, org_id=current_user["org_id"])
    if result.get("status") == "error":
        raise HTTPException(status_code=400, detail=result.get("message"))
    return result


# ============================================================================
# TAKEOFF ENDPOINT (legacy — kept for backward compat)
# ============================================================================

@router.get("/projects/{project_id}/takeoff")
def get_takeoff(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Generate a professional material takeoff from conditions + spec data.
    """
    result = generate_takeoff(project_id, db)
    if result.get("status") == "error":
        raise HTTPException(status_code=400, detail=result.get("message"))
    return result


# ============================================================================
# SAVED ESTIMATES — PERSIST & LOAD TAKEOFF DATA
# ============================================================================

class SaveEstimateRequest(BaseModel):
    estimate_data: dict


@router.post("/projects/{project_id}/save-estimate")
def save_estimate(
    project_id: int,
    body: SaveEstimateRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Save (or update) the takeoff/estimate for a project."""
    import json
    org_id = current_user["org_id"]
    data = body.estimate_data

    grand_total = None
    system_type = None
    roof_area_sf = None
    summary = data.get("summary", {})
    if summary:
        cost_summary = summary.get("cost_summary", {})
        grand_total = cost_summary.get("grand_total")
        system_type = summary.get("system_type")
        roof_area_sf = summary.get("roof_area_sf")

    existing = db.query(SavedEstimate).filter(
        SavedEstimate.project_id == project_id,
        SavedEstimate.org_id == org_id,
    ).first()

    if existing:
        existing.estimate_data = json.dumps(data)
        existing.grand_total = grand_total
        existing.system_type = system_type
        existing.roof_area_sf = roof_area_sf
        existing.version = (existing.version or 1) + 1
        existing.updated_at = datetime.datetime.utcnow()
        db.commit()
        db.refresh(existing)
        return {
            "message": "Estimate updated",
            "id": existing.id,
            "version": existing.version,
            "grand_total": existing.grand_total,
        }
    else:
        saved = SavedEstimate(
            org_id=org_id,
            project_id=project_id,
            version=1,
            estimate_data=json.dumps(data),
            grand_total=grand_total,
            system_type=system_type,
            roof_area_sf=roof_area_sf,
            created_by=current_user["user_id"],
        )
        db.add(saved)
        db.commit()
        db.refresh(saved)
        return {
            "message": "Estimate saved",
            "id": saved.id,
            "version": saved.version,
            "grand_total": saved.grand_total,
        }


@router.get("/projects/{project_id}/saved-estimate")
def get_saved_estimate(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Load the saved estimate for a project (if any)."""
    import json
    org_id = current_user["org_id"]

    saved = db.query(SavedEstimate).filter(
        SavedEstimate.project_id == project_id,
        SavedEstimate.org_id == org_id,
    ).first()

    if not saved:
        return {"saved": False}

    return {
        "saved": True,
        "id": saved.id,
        "version": saved.version,
        "grand_total": saved.grand_total,
        "system_type": saved.system_type,
        "roof_area_sf": saved.roof_area_sf,
        "updated_at": str(saved.updated_at),
        "estimate_data": json.loads(saved.estimate_data),
    }
